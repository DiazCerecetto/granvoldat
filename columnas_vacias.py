import openpyxl

def detect_empty_columns(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    empty_columns = []
    
    for col in range(1, sheet.max_column + 1):
        column_data = [sheet.cell(row=row, column=col).value for row in range(1, sheet.max_row + 1)]
        
        # Verifica si todos los valores en la columna son None o vacíos
        if all(value is None or value == '' for value in column_data):
            empty_columns.append(col)
    
    return empty_columns

file_path = "entrada/CES-2024.xlsx"
empty_columns = detect_empty_columns(file_path)
if empty_columns:
    print("Columnas vacías encontradas en las siguientes columnas:")
    print(empty_columns)
else:
    print("No se encontraron columnas vacías en el archivo.")
