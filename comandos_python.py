
# pip install openpyxl
# pip install pandas
# pip install pyproj
# pip install xlrd
# pip install numpy
import json
import openpyxl
import os
import pandas as pd
from pyproj import Proj

#############################################
################## TESTING ##################
#############################################

def leer_archivo(archivo):
    with open(archivo, 'r') as file:
        return json.load(file)
    
def contar_departamentos(centros_educativos, ubicaciones):
    departamentos = {}
    # el json tiene un campo idubicacion, para esto debemos abrir el json 
    # ubicaciones y buscar el departamento
    for centro_educativo in centros_educativos:
        idubicacion = centro_educativo['idubicacion']
        for ubicacion in ubicaciones:
            if ubicacion['idUbicacion'] == idubicacion:
                departamento = ubicacion['nomdepartamento']
                if departamento not in departamentos:
                    departamentos[departamento] = {}
                    departamentos[departamento]['centros'] = 0
                departamentos[departamento]['centros'] += 1
                break
    # a su vez, contamos cuantas ubicaciones tiene ese departamento, en teoria si todo esta bien,
    # deberia ser igual o menor a la cantidad de centros educativos, es decir,
    # debemos tener al menos un centro por ubicacion
    
    for ubicacion in ubicaciones:
        if ubicacion['nomdepartamento'] not in departamentos:
            departamentos[ubicacion['nomdepartamento']] = {}
            departamentos[ubicacion['nomdepartamento']]['centros'] = 0
        if 'ubicaciones' not in departamentos[ubicacion['nomdepartamento']]:
            departamentos[ubicacion['nomdepartamento']]['ubicaciones'] = 0
        departamentos[ubicacion['nomdepartamento']]['ubicaciones'] += 1
    
    return departamentos

def imprimir_resultados(departamentos):
    for departamento, cantidad in departamentos.items():
        print(f'{departamento}: {cantidad}')

def centros_por_departamento():
    centros_educativos = leer_archivo('salida/r3/centros_educativos_0.json')
    departamentos = leer_archivo('salida/r3/ubicaciones.json')
    slaida = contar_departamentos(centros_educativos,departamentos)
    imprimir_resultados(slaida)
    
def ubicaciones_departamento(nombre_departamento, ubicaciones):
    for ubicacion in ubicaciones:
        if ubicacion['nomdepartamento'] == nombre_departamento:
            print(ubicacion)
        
        
##################################################
###### chequear comas en campos específicos ######
##################################################

def check_commas_in_specific_fields(file_path, field_indices):
    try:
        # Cargar el archivo xlsx
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Iterar sobre cada fila y revisar los campos especificados
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for index in field_indices:
                if index >= len(row):
                    print(f"El índice {index} está fuera del rango de la fila.")
                    continue
                field = row[index]
                if isinstance(field, str) and ',' in field:
                    print(f"Se encontró una coma en el campo {index + 1}:", field)
                    return True
        
        print("No se encontraron comas en los campos especificados de ninguna fila.")
        return False
    
    except FileNotFoundError:
        print("El archivo no fue encontrado.")
        return False
    except Exception as e:
        print("Ocurrió un error:", e)
        return False








###############################################################
################## PREPROCESAMIENTO DE DATOS ##################
###############################################################

def detect_empty_columns(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    empty_columns = []
    
    for col in range(1, sheet.max_column + 1):
        column_data = [sheet.cell(row=row, column=col).value for row in range(1, sheet.max_row + 1)]
        
        # Verifica si todos los valores en la columna son None o vacíos
        if all(value is None or value == '' for value in column_data):
            empty_columns.append(col)
    
    return empty_columns

def utm_to_latlon(x, y):
        # Si x esta entre -53 y -58 e y entre -30 y -35, entonces no hacemos nada y retornamos los numeros tal cual
        # Si por otro lado estos numeros dividido 1000000 estan entre -53 y -58 e -30 y -35, entonces los dividimos por 1000000
        if -53 >= x >= -58 and -30 >= y >= -35:
            return y, x
        elif -53 <= x/1000000 <= -58 and -30 <= y/1000000 <= -35:
            return y/1000000, x/1000000
        utm_proj = Proj(proj='utm', zone=21, south=True, ellps='WGS84', datum='WGS84')
        lon, lat = utm_proj(x, y, inverse=True)
        return lat, lon

# para cada valor de latitud y longitud, cambiar los . por ,
def cambiar_puntos_por_comas(archivo):
    df = pd.read_excel(archivo)
    df['latitud'] = df['latitud'].apply(lambda x: str(x).replace('.', ','))
    df['longitud'] = df['longitud'].apply(lambda x: str(x).replace('.', ','))
    df.to_excel(archivo, index=False)
    
def pasar_a_latlon(ruta_entrada, ruta_salida):
        df = pd.read_excel(ruta_entrada)
        df[['latitud', 'longitud']] = df.apply(lambda row: pd.Series(utm_to_latlon(row['x'], row['y'])), axis=1)
        df['latitud'] = df['latitud'].apply(lambda x: '{:.6f}'.format(x))
        df['longitud'] = df['longitud'].apply(lambda x: '{:.6f}'.format(x))
        df.to_excel(ruta_salida, index=False)
        
def procesar_coordenadas(archivo_escuelas, archivo_liceos, archivo_escuelas_latlon, archivo_liceos_latlon):
    pasar_a_latlon(archivo_escuelas, archivo_escuelas_latlon)
    pasar_a_latlon(archivo_liceos, archivo_liceos_latlon)
    cambiar_puntos_por_comas(archivo_escuelas_latlon)
    cambiar_puntos_por_comas(archivo_liceos_latlon)

def eliminar_tildes_departamentos(file_path):
    #eliminar tildes en la columna Departamento
    df = pd.read_excel(file_path)
    df['Departamento'] = df['Departamento'].str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8')
    df.to_excel(file_path, index=False)
    
###############################################################
################## EJECUCIÓN DE COMANDOS ######################
###############################################################
if __name__ == '__main__':
    # Preguntar por la opcion a ejecutar
    
    print('Opciones:')
    print('1. Testing')
    print('2. Preprocesamiento')
    opcion = input('Ingrese la opcion a ejecutar: ')
    if opcion == '1':
        print('Testing')
        print('Opciones:')
        print('1. Centros por departamento')
        opcion = input('Ingrese la opcion a ejecutar: ')
        if opcion == '1':
            centros_por_departamento()
        else:
            print('Opcion no valida')
    elif opcion == '2':
        opcion = input('Ingrese la opcion a ejecutar: ')
        print('Preprocesamiento')
        print('Opciones:')
        print('1. Procesar coordenadas')
        print('2. Chequear comas en campos específicos')
        print('3. Detectar columnas vacías')
        print('4. Eliminar tildes en departamentos')
        if opcion == '1':
            # Opcion procesar coordenadas
            print('Procesando coordenadas...')
            print('-- Archivos de entrada: entrada/CEIP.xlsx, entrada/CES.xlsx')
            print('-- Archivos de salida: entrada/CEIP-2024.xlsx, entrada/CES-2024.xlsx')
            procesar_coordenadas("entrada/CEIP.xlsx", "entrada/CES.xlsx", "entrada/CEIP-2024.xlsx", "entrada/CES-2024.xlsx")
        elif opcion == '2':
            file_path = "entrada/CES-2024.xlsx"
            fields_to_check = [9, 10, 32] 
            check_commas_in_specific_fields(file_path, fields_to_check)
        elif opcion == '3':
            file_path = "entrada/CES-2024.xlsx"
            empty_columns = detect_empty_columns(file_path)
            if empty_columns:
                print("Columnas vacías encontradas:")
                print(empty_columns)
            else:
                print("No se encontraron columnas vacías en el archivo.")
        elif opcion == '4':
            file_path = "entrada/CEIP-2024.xlsx"
            eliminar_tildes_departamentos(file_path)